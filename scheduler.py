"""
scheduler.py — ส่งงานประจำวัน + Export Status กลับ Excel
รัน: python scheduler.py --push-tasks     # ส่งงานให้ลูกน้องทุกเช้า
     python scheduler.py --export-excel   # Export Status กลับ Excel
     python scheduler.py --daemon         # รันตลอด (cron-like)
"""

import os, sys, time
from datetime import datetime, date
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from dotenv import load_dotenv

load_dotenv()

STATUS_COLORS = {
    "waiting":     "FFFFFF",   # ขาว
    "in_progress": "FFF3CD",   # เหลืองอ่อน
    "qc":          "CCE5FF",   # ฟ้าอ่อน
    "done":        "D4EDDA",   # เขียวอ่อน
    "issue":       "F8D7DA",   # แดงอ่อน
}

STATUS_LABELS = {
    "waiting":     "รอดำเนินการ",
    "in_progress": "กำลังผลิต",
    "qc":          "ตรวจสอบ QC",
    "done":        "เสร็จสิ้น",
    "issue":       "มีปัญหา",
}

# ─── PUSH TASKS ──────────────────────────────────────────────────
def push_daily_tasks():
    """ส่ง Flex Message งานประจำวันให้ลูกน้องทุกคน"""
    # import ที่นี่เพื่อหลีกเลี่ยง circular import
    import line_bot as bot
    print(f"[{now()}] กำลังส่งงานประจำวัน...")
    bot.push_daily_tasks()
    print(f"[{now()}] ส่งงานเสร็จสิ้น")

# ─── EXPORT STATUS → EXCEL ───────────────────────────────────────
def export_status_to_excel(output_path: str = None):
    """
    ดึงข้อมูลจาก Google Sheets → เขียนกลับเป็น Excel
    พร้อม Color-code ตาม Status
    """
    import db_handler

    if output_path is None:
        today = date.today().strftime("%Y-%m-%d")
        output_path = f"production_status_{today}.xlsx"

    print(f"[{now()}] กำลัง Export Status → {output_path}")

    # ดึงข้อมูลจาก Google Sheets
    import gspread
    ss = db_handler.open_spreadsheet()
    ws = ss.worksheet(db_handler.SHEET_PLAN)
    data = ws.get_all_records()
    df = pd.DataFrame(data)

    if df.empty:
        print("ไม่พบข้อมูล")
        return

    # แปลง Status เป็น Label ภาษาไทย
    df["สถานะการผลิต"] = df["production_status"].map(STATUS_LABELS).fillna(df["production_status"])

    # เรียงลำดับคอลัมน์สำหรับรายงาน
    report_cols = [
        "Start Date", "Finish Date", "Machine", "Big/ Small batch",
        "BATCH NO.", "SEFI Saturn Description", "Batch Size (KG)",
        "FG Saturn Description", "Order Qty (KG)",
        "สถานะการผลิต", "updated_by", "updated_at", "issue_note"
    ]
    # กรองเฉพาะคอลัมน์ที่มีอยู่จริง
    cols = [c for c in report_cols if c in df.columns]
    df_report = df[cols].copy()

    # เขียนไฟล์ Excel
    wb = openpyxl.Workbook()
    ws_out = wb.active
    ws_out.title = "Production Status"

    # Header row
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    for col_idx, col_name in enumerate(df_report.columns, start=1):
        cell = ws_out.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows + color by status
    for row_idx, row in df_report.iterrows():
        status_key = df.at[row_idx, "production_status"] if "production_status" in df.columns else "waiting"
        fill_color = STATUS_COLORS.get(status_key, "FFFFFF")
        row_fill   = PatternFill("solid", fgColor=fill_color)

        for col_idx, value in enumerate(row, start=1):
            cell       = ws_out.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.fill  = row_fill
            cell.font  = Font(name="Arial", size=10)

    # Auto column width
    for col in ws_out.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Sheet 2: Summary
    ws_sum = wb.create_sheet("Summary")
    summary = db_handler.get_daily_summary()
    ws_sum["A1"] = "สรุปสถานะการผลิต"
    ws_sum["A1"].font = Font(bold=True, size=14, name="Arial")
    ws_sum["A2"] = f"วันที่: {summary['date']}"

    row_s = 4
    for key, label in STATUS_LABELS.items():
        count = summary.get(key, 0)
        ws_sum.cell(row=row_s, column=1, value=label).font = Font(name="Arial")
        cell = ws_sum.cell(row=row_s, column=2, value=count)
        cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[key])
        cell.font = Font(name="Arial")
        row_s += 1

    ws_sum.cell(row=row_s+1, column=1, value="รวมทั้งหมด").font = Font(bold=True, name="Arial")
    ws_sum.cell(row=row_s+1, column=2, value=summary["total"]).font = Font(bold=True, name="Arial")

    wb.save(output_path)
    print(f"[{now()}] ✅ Export เสร็จสิ้น → {output_path}")
    return output_path

# ─── DAEMON (รันตลอด) ────────────────────────────────────────────
def run_daemon():
    """
    รันเป็น background process
    - 07:00 น. → ส่งงานประจำวันให้ลูกน้อง
    - ทุก 30 นาที → Export Status กลับ Excel (เก็บ backup)
    """
    print(f"[{now()}] Scheduler daemon เริ่มทำงาน")
    last_push_date   = None
    last_export_hour = None

    while True:
        now_dt   = datetime.now()
        now_date = now_dt.date()
        now_hour = now_dt.hour
        now_min  = now_dt.minute

        # ส่งงานทุกเช้า 07:00
        if now_hour == 7 and now_min < 5 and last_push_date != now_date:
            push_daily_tasks()
            last_push_date = now_date

        # Export ทุกๆ 30 นาที (07:00-22:00)
        if 7 <= now_hour <= 22 and now_min in (0, 30) and last_export_hour != (now_hour, now_min):
            export_status_to_excel()
            last_export_hour = (now_hour, now_min)

        time.sleep(60)  # ตรวจทุก 1 นาที

def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ─── CLI ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python scheduler.py --push-tasks         ส่งงานวันนี้ทันที")
        print("  python scheduler.py --export-excel       Export Status → Excel")
        print("  python scheduler.py --daemon             รันตลอด (background)")
        sys.exit(0)

    cmd = sys.argv[1]
    if cmd == "--push-tasks":
        push_daily_tasks()
    elif cmd == "--export-excel":
        out = sys.argv[2] if len(sys.argv) > 2 else None
        export_status_to_excel(out)
    elif cmd == "--daemon":
        run_daemon()
    else:
        print(f"ไม่รู้จักคำสั่ง: {cmd}")
